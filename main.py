import datetime
import os
from openpyxl import Workbook, load_workbook
import pandas as pd
import numpy as np
from sklearn import preprocessing
import math
import time
import random
import copy
from sklearn.base import BaseEstimator, ClassifierMixin
from sklearn.svm import SVR
from sklearn.pipeline import make_pipeline
from sklearn.preprocessing import StandardScaler
from SGTM import GTM
from sklearn.model_selection import KFold, train_test_split

from sklearn.metrics import explained_variance_score
from sklearn.metrics import max_error
from sklearn.metrics import mean_absolute_error
from sklearn.metrics import mean_squared_error
from sklearn.metrics import mean_absolute_percentage_error
from sklearn.metrics import median_absolute_error
from sklearn.metrics import r2_score


def root_mean_squared_error(y_true, y_pred):
    return np.sqrt(((np.array(y_pred) - np.array(y_true)) ** 2).mean())


# ===================== допоміжні функції ==================

def normalize(df):
    x = df.values 
    max_abs_scaler = preprocessing.StandardScaler()
    x_scaled = max_abs_scaler.fit_transform(x)
    return pd.DataFrame(x_scaled)

# ============================================================


class GRNN(BaseEstimator, ClassifierMixin):
    def __init__(self, name = "GRNN", sigma = 0.1):
        self.name = name
        self.sigma = 2 * np.power(sigma, 2)
        
    def predict(self, instance_X, trn_X, trn_y):
        gausian_distances = np.exp(-np.power(np.sqrt((np.square(trn_X-instance_X).sum(axis=1))),2)\
                                   / self.sigma)
        gausian_distances_sum = gausian_distances.sum()
        if gausian_distances_sum < math.pow(10, -7): gausian_distances_sum = math.pow(10, -7)
        result = np.multiply(gausian_distances, trn_y).sum() / gausian_distances_sum
        return result


def generate_noise(dataset,start=-0.02,end=0.02,kmx=4):
    import copy
    df = copy.deepcopy(dataset.values)
    res = []
    for i in range(len(df)):
        for _ in range(kmx):
            buf = copy.deepcopy(df[i])
            for k in range(len(buf)):
                buf[k] = random.uniform(start, end)+buf[k]
            res.append(buf)
    return res


def generate_noise_values(dataset,start=-0.02,end=0.02,kmx=4):
    import copy
    df = copy.deepcopy(dataset.values)
    res = []
    for i in range(len(df)):
        for _ in range(kmx):
            res.append(random.uniform(start, end)+df[i])
    #return pd.DataFrame(res)
    return res


def write_to_excel(data, file_name):
    # Create a timestamp string
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    # Add the timestamp to the file name
    file_name_with_timestamp = f"{file_name}_{timestamp}.xlsx"
    # Convert the data to a DataFrame
    df = pd.DataFrame(data)
    # Write the DataFrame to Excel

    df.to_excel(file_name_with_timestamp, index=False)


def write_errors_to_excel(pred,y,mode,kmax,noiseRange,name='train', ):

    global dataset

    if os.path.exists(f'{dataset}_{name}_mode-{mode}.xlsx.xlsx'):
        pass
    else:
        # Create new Excel workbook
        workbook = Workbook()

        try:
            workbook.create_sheet('Sheet')
        except ValueError:
            pass

        excel_row = [
            'mode',
            'kmax',
            'noiseRange',
            'explained_variance_score_value',
            'root_mean_squared_error_value',
            'r2_score_value',
            'median_absolute_error_value',
            'mean_absolute_percentage_error_value',
            'mean_squared_error_value',
            'mean_absolute_error_value',
            'max_error_value'
        ]

        worksheet = workbook['Sheet']
        worksheet.append(excel_row)
        workbook.save(filename=f'{dataset}_{name}_mode-{mode}.xlsx.xlsx')

    explained_variance_score_value = explained_variance_score(y, pred)
    max_error_value = max_error(y, pred)
    mean_absolute_error_value = mean_absolute_error(y, pred)
    mean_squared_error_value = mean_squared_error(y, pred)
    mean_absolute_percentage_error_value = mean_absolute_percentage_error(y, pred)
    median_absolute_error_value = median_absolute_error(y, pred)
    r2_score_value = r2_score(y, pred)
    root_mean_squared_error_value = root_mean_squared_error(y, pred)

    print('='*40)
    print()
    print(f'explained_variance_score_value : {explained_variance_score_value}')
    print(f'root_mean_squared_error_value : {root_mean_squared_error_value}')
    print(f'r2_score_value : {r2_score_value}')
    print(f'median_absolute_error_value : {median_absolute_error_value}')
    print(f'mean_absolute_percentage_error_value : {mean_absolute_percentage_error_value}')
    print(f'mean_squared_error_value : {mean_squared_error_value}')
    print(f'mean_absolute_error_value : {mean_absolute_error_value}')
    print(f'max_error_value : {max_error_value}')
    print()

    excel_row = [
        mode,
        kmax,
        noiseRange,
        explained_variance_score_value,
        root_mean_squared_error_value,
        r2_score_value,
        median_absolute_error_value,
        mean_absolute_percentage_error_value,
        mean_squared_error_value,
        mean_absolute_error_value,
        max_error_value
    ]
    
    workbook = load_workbook(filename=f'{dataset}_{name}_mode-{mode}.xlsx.xlsx')
    
    worksheet = workbook['Sheet']
    worksheet.append(excel_row)
    workbook.save(filename=f'{dataset}_{name}_mode-{mode}.xlsx.xlsx')
    worksheet.append([])


def find_best_sigma(train_X,train_y,test_X,test_y):
    grnn_training_error=1
    val = 10000000
    print('started selecting best sigma for GRNN')
    start_time = time.time()
    for s in np.arange(0.001, 1, 0.001):

        grnn = GRNN(sigma=s)
        predictions = np.apply_along_axis(lambda i: grnn.predict(i, train_X, train_y), axis=1, arr=test_X)
        
        grnn_training_error = mean_absolute_percentage_error(test_y, predictions)
        if grnn_training_error<val:
            val=grnn_training_error
            best_sigma=s
    
    print()
    print(f'training time : {time.time() - start_time}')
    print(f'best sigma : {best_sigma}')
    print(f'MAPE error when best sigma : {grnn_training_error}')

    return best_sigma


def create_pred_dataset(train_X, train_y, noiseRange, mode):

    noise = np.random.uniform(-noiseRange, noiseRange, size=(kmax,len(train_X.values[0])))

    predictions = []
    
    for k in range(kmax):
        # generate noise array for this row of train_X
        
        # add noise to this row of train_X
        train_X_noised = train_X * noise[k] + train_X
        train_y_noised = train_y

        preds = []

        for i in range(len(train_X_noised)):
            grnn = GRNN(sigma=best_sigma)    

            train_X_noised_sliced = train_X_noised.drop(i).copy()
            train_y_noised_sliced = train_y_noised.drop(i).copy()

            pred = grnn.predict(train_X_noised.values[i],
                                train_X_noised_sliced,
                                train_y_noised_sliced)

            preds.append(pred)
        
        predictions.append(preds)


    to_excel = [item for sublist in predictions for item in sublist]

    write_errors_to_excel(np.tile(train_y_noised,kmax), to_excel, mode,kmax, noiseRange,'train')

    to_convert = {f"pred_{i}": col for i, col in enumerate(predictions)}

    pred_dataset = pd.DataFrame(to_convert)

    return pred_dataset



def worker (train, test, use, mode=1, kmax=4, noiseRange=0.02, model=None ):
    global best_sigma

    train_X = train.iloc[:, :-1]
    train_y = train.iloc[:,-1]
    test_X = test.iloc[:, :-1]
    test_y = test.iloc[:,-1]

    train_X = normalize(train_X)
    test_X = normalize(test_X)

    if best_sigma <= 0:
        best_sigma = find_best_sigma(train_X,train_y,test_X,test_y)
    

    pred_dataset = create_pred_dataset(train_X, train_y, noiseRange, mode)

    print('creating extended vectors , step 4.4')

    if mode == 1:
        vector = pd.concat([train_X, pred_dataset], axis=1, join="inner")
    if mode == 2:
        vector = pd.concat([train_y, pred_dataset], axis=1, join="inner")

    # write_to_excel(vector,'vector')
    

    if not model:
        model = make_pipeline(StandardScaler(), SVR(C=kmax, epsilon=0.2,kernel='rbf'))

    min_len = min(len(vector), len(train_y))

    vector = vector.iloc[:min_len]
    train_y = train_y.iloc[:min_len]

    # vector = normalize(vector)

    model.fit(vector, train_y)
    
    normalized_use = normalize(use)

    use_X = normalized_use.iloc[:, :-1]
    use_y = use.iloc[:,-1]

    use_X_noised = generate_noise(use_X, -noiseRange, noiseRange, kmax)
    # use_y_noised = generate_noise_values(use_y, -noiseRange, noiseRange, kmax)
    # use_y_noised = [elem for elem in use_y for _ in range(kmax)]
    

    predictions = []
    buf=[]
    for i in range(len(use_X_noised)):
        grnn = GRNN(sigma=best_sigma)

        # use_X_noised_sliced = use_X_noised[:i] + use_X_noised[i+1:]
        # use_y_noised_sliced = use_y_noised[:i] + use_y_noised[i+1:]

        pred = grnn.predict(use_X_noised[i],
                                train_X,
                                train_y)
        buf.append(pred)
        if len(buf)==kmax:
            predictions.append(buf)
            buf=[]  
    
    result = pd.DataFrame(predictions)
    

    if mode == 1:
        use_vector = pd.concat([pd.DataFrame(use_X), pd.DataFrame(result)], axis=1, join="inner")

    if mode == 2:
        use_vector = pd.concat([pd.DataFrame(use_y), pd.DataFrame(result)], axis=1, join="inner")

    predictions = model.predict(use_vector)


    print('best_sigma : ', best_sigma)
    print()

    write_errors_to_excel(predictions, use_y,mode,kmax, noiseRange,'test')



def get_cross_validation_datasets(train, test, n_splits=10, random_state=None):
    kf = KFold(n_splits=n_splits, shuffle=True, random_state=random_state)

    mixed_dataset = pd.concat([train, test], axis=0, ignore_index=True)
    datasets = []
    for train_idx, test_idx in kf.split(mixed_dataset):
        a = mixed_dataset.iloc[train_idx]
        b = mixed_dataset.iloc[test_idx]

        datasets.append([a,b])
    return datasets


if __name__ == '__main__':
    
    model = None

    datasets = ['OBD']

    # kmax = 15

    for dataset in datasets:
        best_sigma = 0.303
        for kmax in range (1,20,1):

            buf = pd.read_csv(f"datasets/{dataset}Train.csv", header=None)
            
            train_buf, test_buf = train_test_split(buf, test_size=0.2, random_state=42)

            # Convert train_buf and test_buf back to pandas DataFrames
            train = pd.DataFrame(train_buf).reset_index(drop=True)
            test = pd.DataFrame(test_buf).reset_index(drop=True)
            

            # datasets = get_cross_validation_datasets(train,test)

            use = pd.read_csv(f"datasets/{dataset}Test.csv", header=None)

            # for train, test in datasets:

            worker(
                train, 
                test,
                use,
                mode=1,
                kmax=kmax,
                noiseRange=0.1)

